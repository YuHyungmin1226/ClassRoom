from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash, send_from_directory, send_file, abort, current_app
from werkzeug.utils import secure_filename
from PIL import Image
import os
import uuid
import io
import zipfile
from datetime import datetime
from pathlib import PurePosixPath
from . import db
from .models import Admin, ClassGroup, Session, Flag, QuizQuestion, QuizResponse, Upload

main = Blueprint('main', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'txt', 'docx', 'mp4', 'webm', 'mov'}
CLASS_TYPES = frozenset({'classmap', 'classwrite', 'classdraw', 'classquiz'})

WORKSPACE_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
GAME_LIBRARY = {
    'dialike': {
        'id': 'dialike',
        'name': 'DIALIKE',
        'subtitle': 'Dark Action RPG',
        'description': '디아블로 스타일 전투, 파밍, 장비 성장을 담은 웹 액션 RPG',
        'input': '키보드 · 마우스',
        'folder': os.path.abspath(os.environ.get(
            'CLASSGAME_DIALIKE_PATH',
            os.path.join(WORKSPACE_ROOT, 'dialike'),
        )),
        'entry': 'index.html',
        'cover': 'assets/tile_stone.png',
        'public_files': frozenset({'index.html', 'style.css'}),
        'public_dirs': {
            'assets': frozenset({'.png', '.jpg', '.jpeg', '.gif', '.webp'}),
            'js': frozenset({'.js'}),
        },
    },
}


def _available_games():
    games = []
    for game in GAME_LIBRARY.values():
        item = dict(game)
        item['available'] = os.path.isfile(os.path.join(game['folder'], game['entry']))
        item['play_url'] = url_for('main.classgame_files', game_id=game['id'])
        item['cover_url'] = (
            url_for('main.classgame_files', game_id=game['id'], filename=game['cover'])
            if (
                _public_game_filename(game, game.get('cover'))
                and os.path.isfile(os.path.join(game['folder'], game['cover']))
            ) else None
        )
        games.append(item)
    return games


def _get_game_or_404(game_id):
    game = GAME_LIBRARY.get(game_id)
    if not game or not os.path.isfile(os.path.join(game['folder'], game['entry'])):
        abort(404)
    return game


def _public_game_filename(game, filename):
    """Return a normalized public asset path, or None for private files."""
    if not filename:
        return None

    path = PurePosixPath(str(filename).replace('\\', '/'))
    parts = path.parts
    if (
        not parts
        or path.is_absolute()
        or '..' in parts
        or any(part.startswith('.') for part in parts)
    ):
        return None

    normalized = path.as_posix()
    if normalized in game.get('public_files', ()):
        return normalized

    allowed_extensions = game.get('public_dirs', {}).get(parts[0])
    if len(parts) > 1 and allowed_extensions and path.suffix.lower() in allowed_extensions:
        return normalized
    return None

@main.route('/uploads/<filename>')
def uploaded_file(filename):
    relative_path = f"uploads/{filename}"
    upload = Upload.query.filter(
        (Upload.file_path == relative_path) | (Upload.thumbnail_path == relative_path)
    ).first()
    legacy_flag = None
    if not upload:
        legacy_flag = Flag.query.filter(
            (Flag.file_path == relative_path) | (Flag.thumbnail_path == relative_path)
        ).first()
        if not legacy_flag:
            abort(404)

    # Closed-session attachments remain available to administrators only.
    if not session.get('admin_logged_in'):
        owning_session = upload.session if upload else legacy_flag.session
        if not owning_session.is_open:
            return "This session is closed.", 403
    extension = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return send_from_directory(
        current_app.config['UPLOAD_FOLDER'],
        filename,
        as_attachment=extension in {'html', 'htm', 'svg', 'xml'},
    )

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def delete_file_safe(file_path):
    if not file_path:
        return
    filename = os.path.basename(file_path)
    abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    try:
        if os.path.exists(abs_path) and os.path.isfile(abs_path):
            os.remove(abs_path)
    except Exception as e:
        print(f"Failed to delete orphaned file {abs_path}: {e}")


def delete_file_if_unreferenced(file_path):
    if not file_path:
        return
    referenced = Upload.query.filter(
        (Upload.file_path == file_path) | (Upload.thumbnail_path == file_path)
    ).first() or Flag.query.filter(
        (Flag.file_path == file_path) | (Flag.thumbnail_path == file_path)
    ).first()
    if not referenced:
        delete_file_safe(file_path)


def delete_upload_records_before_flags(uploads):
    uploads = list(uploads)
    for upload in uploads:
        upload.flag_id = None
    if uploads:
        db.session.flush()

# --- Admin Routes ---
@main.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        admin = Admin.query.first()
        if password and admin and admin.check_password(password):
            session['admin_logged_in'] = True
            return redirect(url_for('main.admin_classroom'))
        flash('Invalid password')
    return render_template('admin_login.html')

@main.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('main.admin_login'))

@main.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    class_type = request.args.get('type', 'classmap')
    active_classes = ClassGroup.query.filter_by(is_active=True, class_type=class_type).all()
    past_classes = ClassGroup.query.filter_by(is_active=False, class_type=class_type).all()
    return render_template('admin_dashboard.html', 
                           active_classes=active_classes, 
                           past_classes=past_classes,
                           current_type=class_type)


@main.route('/admin/create_class', methods=['POST'])
def create_class():
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    name = request.form.get('name')
    class_type = request.form.get('class_type', 'classmap')
    if class_type not in CLASS_TYPES:
        abort(400, description='Invalid class type.')
    
    if not name:
        name = f"Class {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    new_class = ClassGroup(name=name, class_type=class_type)
    db.session.add(new_class)
    db.session.commit()
    return redirect(url_for('main.admin_dashboard', type=class_type))

@main.route('/admin/close_class/<class_id>', methods=['POST'])
def close_class(class_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    c = db.get_or_404(ClassGroup, class_id)
    c.is_active = False
    # 세션까지 함께 닫아야 한다 — 그렇지 않으면 클래스는 닫혔어도 자식 세션은 is_active=True로
    # 남아, view_session/소켓 쓰기/첨부파일 다운로드가 여전히 세션의 is_active만 검사하므로
    # 비관리자가 "닫힌" 클래스의 콘텐츠에 그대로 접근/쓰기할 수 있게 된다.
    for s in c.sessions:
        s.is_active = False
    db.session.commit()
    return redirect(url_for('main.admin_dashboard', type=c.class_type))

@main.route('/admin/delete_class/<class_id>', methods=['POST'])
def delete_class(class_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    c = db.get_or_404(ClassGroup, class_id)
    class_type = c.class_type
    class_name = c.name  # 삭제 후 접근 방지 위해 미리 캡처

    files_to_delete = set()
    uploads_to_delete = []
    for s in c.sessions:
        for upload in s.uploads:
            files_to_delete.update((upload.file_path, upload.thumbnail_path))
            uploads_to_delete.append(upload)
        for f in s.flags:
            files_to_delete.update((f.file_path, f.thumbnail_path))

    delete_upload_records_before_flags(uploads_to_delete)
    db.session.delete(c)
    db.session.commit()
    for file_path in files_to_delete:
        delete_file_if_unreferenced(file_path)
    flash(f"Class '{class_name}' has been successfully deleted.")
    return redirect(url_for('main.admin_dashboard', type=class_type))

@main.route('/admin/class/<class_id>')
def admin_class(class_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    c = db.get_or_404(ClassGroup, class_id)
    active_sessions = Session.query.filter_by(class_id=class_id, is_active=True).all()
    past_sessions = Session.query.filter_by(class_id=class_id, is_active=False).all()
    return render_template('admin_class.html', class_group=c, active_sessions=active_sessions, past_sessions=past_sessions)

@main.route('/admin/class/<class_id>/create_session', methods=['POST'])
def create_session(class_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    c = db.get_or_404(ClassGroup, class_id)
    if not c.is_active:
        abort(409, description='Cannot create a session in a closed class.')
    name = request.form.get('name')
    if not name:
        name = f"Session {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    new_session = Session(name=name, class_id=c.id)
    db.session.add(new_session)
    db.session.commit()
    return redirect(url_for('main.admin_class', class_id=c.id))

@main.route('/admin/close_session/<session_id>', methods=['POST'])
def close_session(session_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    s = db.get_or_404(Session, session_id)
    s.is_active = False
    db.session.commit()
    return redirect(url_for('main.admin_class', class_id=s.class_id))

@main.route('/admin/delete_session/<session_id>', methods=['POST'])
def delete_session(session_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    s = db.get_or_404(Session, session_id)
    class_id = s.class_id
    session_name = s.name  # 삭제 후 접근 방지 위해 미리 캡처

    files_to_delete = set()
    for upload in s.uploads:
        files_to_delete.update((upload.file_path, upload.thumbnail_path))
    for f in s.flags:
        files_to_delete.update((f.file_path, f.thumbnail_path))

    delete_upload_records_before_flags(s.uploads)
    db.session.delete(s)
    db.session.commit()
    for file_path in files_to_delete:
        delete_file_if_unreferenced(file_path)
    flash(f"Session '{session_name}' has been successfully deleted.")
    return redirect(url_for('main.admin_class', class_id=class_id))

@main.route('/admin/class/<class_id>/quiz_results')
def admin_class_quiz_results(class_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    c = db.get_or_404(ClassGroup, class_id)
    sessions = Session.query.filter_by(class_id=class_id).all()
    return render_template('class_quiz_results.html', class_group=c, sessions=sessions)

@main.route('/admin/settings')
def admin_settings():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    return render_template('admin_settings.html')

@main.route('/admin/change_password', methods=['POST'])
def change_password():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    admin = Admin.query.first()

    # 현재 비밀번호 확인 (CSRF/세션 탈취 시 비밀번호 변경 악용 방지)
    if not admin or not current_password or not admin.check_password(current_password):
        flash('현재 비밀번호가 올바르지 않습니다.')
        return redirect(url_for('main.admin_settings'))

    if not new_password or len(new_password) < 10:
        flash('새 비밀번호는 10자 이상이어야 합니다.')
        return redirect(url_for('main.admin_settings'))

    admin.set_password(new_password)
    db.session.commit()
    flash('Password changed successfully.')
    return redirect(url_for('main.admin_settings'))

@main.route('/admin/reset_data', methods=['POST'])
def reset_data():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    import shutil
    # 벌크 delete()는 ORM cascade를 타지 않으므로, 자식 테이블부터 명시적으로 모두 삭제한다.
    # (이전 코드는 QuizQuestion/QuizResponse를 누락해 고아 레코드가 남았음)
    db.session.query(Upload).delete()
    db.session.query(QuizResponse).delete()
    db.session.query(QuizQuestion).delete()
    db.session.query(Flag).delete()
    db.session.query(Session).delete()
    db.session.query(ClassGroup).delete()
    db.session.commit()
    
    upload_folder = current_app.config['UPLOAD_FOLDER']
    if os.path.exists(upload_folder):
        for filename in os.listdir(upload_folder):
            file_path = os.path.join(upload_folder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))
                
    flash('All data has been successfully reset.')
    return redirect(url_for('main.admin_settings'))

@main.route('/admin/export_markdown')
def export_markdown():
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    # Create a ZIP file in memory
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Fetch all classes, sessions, and flags
        classes = ClassGroup.query.all()
        
        for c in classes:
            class_folder = secure_filename(c.name) or f"Class_{c.id[:8]}"
            
            for s in c.sessions:
                session_folder = secure_filename(s.name) or f"Session_{s.id[:8]}"
                
                for f in s.flags:
                    # Create markdown content
                    created_str = f.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    safe_author = secure_filename(f.author_name) or "Participant"
                    filename = f"{f.id}_{safe_author}_{f.created_at.strftime('%Y%m%d_%H%M%S')}.md"
                    filepath = os.path.join(class_folder, session_folder, filename)
                    
                    md_content = f"""# Post by {f.author_name}
**Date:** {created_str}
**Class:** {c.name}
**Session:** {s.name}
"""
                    if f.x is not None and f.y is not None:
                        md_content += f"**Location:** ({f.x}, {f.y})\n"
                    
                    md_content += "\n---\n\n"
                    md_content += f.text_content if f.text_content else "*No text content*"
                    md_content += "\n\n---\n"
                    
                    if f.file_path:
                        md_content += f"**Attached File:** {f.file_path}\n"
                        attachment_name = os.path.basename(f.file_path)
                        attachment_path = os.path.join(current_app.config['UPLOAD_FOLDER'], attachment_name)
                        if os.path.isfile(attachment_path):
                            attachment_zip_path = os.path.join(class_folder, session_folder, 'attachments', attachment_name)
                            zf.write(attachment_path, attachment_zip_path)
                    
                    # Add to zip
                    zf.writestr(filepath, md_content)
        
        # If no flags found, add a placeholder
        if not Flag.query.first():
            zf.writestr("empty_export.txt", "No posts found to export.")

    memory_file.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        memory_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"classroom_export_{timestamp}.zip"
    )

# --- Portal and Common Routes ---
@main.route('/')
def index():
    # Integrated portal for everyone
    return render_template('classroom_select.html', is_admin=session.get('admin_logged_in', False))

@main.route('/classgame')
def classgame():
    return render_template('class_game.html', games=_available_games())

@main.route('/classgame/<game_id>/files/')
@main.route('/classgame/<game_id>/files/<path:filename>')
def classgame_files(game_id, filename=None):
    game = _get_game_or_404(game_id)
    filename = filename or game['entry']
    public_filename = _public_game_filename(game, filename)
    if not public_filename:
        abort(404)

    root = os.path.realpath(game['folder'])
    target = os.path.realpath(os.path.join(root, *PurePosixPath(public_filename).parts))
    try:
        if os.path.commonpath((root, target)) != root or not os.path.isfile(target):
            abort(404)
    except ValueError:
        abort(404)
    return send_from_directory(root, public_filename)

@main.route('/admin/classroom')
def admin_classroom():
    # Still keep this for explicit admin access
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    return render_template('classroom_select.html', is_admin=True)

# --- Participant Routes ---
@main.route('/classes')
def classes():
    # Participant dashboard showing filtered active classes
    class_type = request.args.get('type', 'classmap')
    active_classes = ClassGroup.query.filter_by(is_active=True, class_type=class_type).all()
    is_admin = session.get('admin_logged_in', False)
    return render_template('index.html', classes=active_classes, current_type=class_type, is_admin=is_admin)

@main.route('/class/<class_id>')
def view_class(class_id):
    c = db.get_or_404(ClassGroup, class_id)
    if not c.is_active and not session.get('admin_logged_in'):
        return "This class is closed.", 403
    active_sessions = Session.query.filter_by(class_id=class_id, is_active=True).all()
    is_admin = session.get('admin_logged_in', False)
    return render_template('class_sessions.html', class_group=c, sessions=active_sessions, is_admin=is_admin)

@main.route('/session/<session_id>')
def view_session(session_id):
    s = db.get_or_404(Session, session_id)
    if not s.is_open and not session.get('admin_logged_in'):
        return "This session is closed.", 403

    is_admin = session.get('admin_logged_in', False)

    # 위조 불가능한 서버측 참여자 식별자 발급 (서명된 Flask 세션에 저장).
    # 게시물 소유권 검증의 신뢰 기준이 되며, 클라이언트가 보낸 client_id는 신뢰하지 않는다.
    if not session.get('participant_id'):
        session['participant_id'] = uuid.uuid4().hex
        session.permanent = True
    participant_id = session['participant_id']

    if s.class_group.class_type == 'classquiz':
        return render_template('quiz_session.html', quiz_session=s,
                               is_admin=is_admin, participant_id=participant_id)
    return render_template('map_session.html', map_session=s,
                           is_admin=is_admin, participant_id=participant_id)

@main.route('/admin/session/<session_id>/export_quiz')
def export_quiz_excel(session_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('main.admin_login'))
    
    from openpyxl import Workbook

    s = db.get_or_404(Session, session_id)
    questions = QuizQuestion.query.filter_by(session_id=session_id).order_by(QuizQuestion.index).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Questions"
    
    # Headers
    headers = ['Idx', 'Type', 'Question', 'Options (split by |)', 'Correct Answer']
    ws.append(headers)
    
    # Data
    for q in questions:
        ws.append([q.index, q.q_type, q.question, q.options, q.correct_answer])
    
    # Save to memory
    memory_file = io.BytesIO()
    wb.save(memory_file)
    memory_file.seek(0)
    
    filename = f"quiz_{secure_filename(s.name)}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        memory_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@main.route('/admin/session/<session_id>/import_quiz', methods=['POST'])
def import_quiz_excel(session_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401

    quiz_session = db.get_or_404(Session, session_id)
    if quiz_session.class_group.class_type != 'classquiz':
        abort(400, description='Quiz imports require a ClassQuiz session.')
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(file)
        ws = wb.active
        imported_questions = []
        quiz_uploads = Upload.query.filter_by(
            session_id=session_id,
            purpose='quiz',
        ).all()
        files_to_delete = {
            path
            for upload in quiz_uploads
            for path in (upload.file_path, upload.thumbnail_path)
            if path
        }

        # 기존 문제를 새 파일로 "동기화"(교체)한다.
        # 교체될 문제에 달린 응답이 고아로 남지 않도록 먼저 삭제한다.
        for upload in quiz_uploads:
            upload.question_id = None
        if quiz_uploads:
            db.session.flush()
        for upload in quiz_uploads:
            db.session.delete(upload)
        if quiz_uploads:
            db.session.flush()
        QuizResponse.query.filter_by(session_id=session_id).delete()
        QuizQuestion.query.filter_by(session_id=session_id).delete()

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3 or not row[2]:
                continue  # Skip if no question text

            q_type = str(row[1]).strip().lower() if row[1] else 'choice'
            if q_type not in {'choice', 'short', 'long'}:
                raise ValueError(f"Invalid question type: {q_type}")

            options = str(row[3]) if len(row) > 3 and row[3] else ""
            correct_answer = str(row[4]) if len(row) > 4 and row[4] else ""
            if q_type == 'choice' and len([opt for opt in options.split('|') if opt.strip()]) < 2:
                raise ValueError(f"Choice question requires at least two options: {row[2]}")
            if q_type in {'choice', 'short'} and not correct_answer:
                raise ValueError(f"{q_type} question requires a correct answer: {row[2]}")

            new_q = QuizQuestion(
                session_id=session_id,
                index=row[0] if row[0] is not None else 0,
                q_type=q_type,
                question=str(row[2]),
                options=options,
                correct_answer=correct_answer
            )
            imported_questions.append(new_q)

        # 유효한 문제가 하나도 없으면 기존 문제/응답을 지운 채로 커밋되지 않도록 중단한다
        # (잘못된 시트/헤더로 업로드 시 전체 데이터가 조용히 삭제되는 것을 방지).
        if not imported_questions:
            raise ValueError('No valid questions found in file; import aborted.')

        for q in imported_questions:
            db.session.add(q)

        db.session.commit()
        for file_path in files_to_delete:
            delete_file_if_unreferenced(file_path)
        return jsonify({'success': True})
    except (ValueError, TypeError, zipfile.BadZipFile, InvalidFileException) as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@main.route('/upload', methods=['POST'])
def upload_file():
    session_id = request.form.get('session_id')
    purpose = request.form.get('purpose', 'flag')
    target_session = db.session.get(Session, session_id) if session_id else None
    if not target_session:
        return jsonify({'error': 'A valid session is required'}), 400

    is_admin = session.get('admin_logged_in', False)
    owner_id = session.get('participant_id') or ('admin' if is_admin else None)
    if purpose == 'quiz':
        if not is_admin or target_session.class_group.class_type != 'classquiz':
            return jsonify({'error': 'Unauthorized upload purpose'}), 403
    elif purpose == 'flag':
        if target_session.class_group.class_type == 'classquiz':
            return jsonify({'error': 'Invalid upload purpose for this session'}), 400
        if not is_admin and not target_session.is_open:
            return jsonify({'error': 'This session is closed'}), 403
        if not owner_id:
            return jsonify({'error': 'Open the session before uploading'}), 403
    else:
        return jsonify({'error': 'Invalid upload purpose'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        extension = filename.rsplit('.', 1)[1].lower()
        if purpose == 'quiz' and extension not in {'png', 'jpg', 'jpeg', 'gif'}:
            return jsonify({'error': 'Quiz uploads must be images'}), 400
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        upload_folder = current_app.config['UPLOAD_FOLDER']
        filepath = os.path.join(upload_folder, unique_filename)
        
        # Ensure upload dir exists
        os.makedirs(upload_folder, exist_ok=True)
        file.save(filepath)
        
        # Create thumbnail if image
        thumbnail_path = None
        if extension in {'png', 'jpg', 'jpeg', 'gif'}:
            thumb_filename = f"thumb_{unique_filename}"
            thumb_filepath = os.path.join(upload_folder, thumb_filename)
            try:
                with Image.open(filepath) as img:
                    img.verify()
                with Image.open(filepath) as img:
                    img.thumbnail((150, 150))
                    img.save(thumb_filepath)
                thumbnail_path = f"uploads/{thumb_filename}"
            except Exception as e:
                if os.path.isfile(thumb_filepath):
                    os.remove(thumb_filepath)
                if purpose == 'quiz':
                    delete_file_safe(f"uploads/{unique_filename}")
                    return jsonify({'error': 'Invalid image file'}), 400
                print(f"Error creating thumbnail: {e}")
                
        file_path = f"uploads/{unique_filename}"
        upload = Upload(
            session_id=target_session.id,
            owner_id=owner_id,
            purpose=purpose,
            file_path=file_path,
            thumbnail_path=thumbnail_path,
        )
        db.session.add(upload)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            delete_file_safe(file_path)
            delete_file_safe(thumbnail_path)
            return jsonify({'error': 'Could not register upload'}), 500

        return jsonify({
            'success': True,
            'file_path': file_path,
            'thumbnail_path': thumbnail_path
        })
    return jsonify({'error': 'Invalid file type'}), 400
